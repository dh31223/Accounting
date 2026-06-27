"""Excel 导出模块。

将交易数据导出为 .xlsx 表格，便于阅读和统计。
"""

from datetime import date
from typing import Optional
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from db.connection import get_connection, close_connection


class ExportService:
    """Excel 导出服务。"""

    ATTR_LABELS = {
        "worthy": "💚 值得",
        "joy": "💛 悦己",
        "necessity": "📋 刚需",
        "waste": "🚫 浪费",
    }
    TYPE_LABELS = {"income": "收入", "expense": "支出"}

    @staticmethod
    def export_transactions(
        filepath: str,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        conn=None,
    ) -> int:
        """导出交易记录到 Excel 文件。

        Args:
            filepath: 目标 .xlsx 文件路径
            start_date: 起始日期（可选）
            end_date: 结束日期（可选）

        Returns:
            导出的记录数
        """
        own_conn = conn is None
        if own_conn:
            conn = get_connection()

        # 查询数据
        query = "SELECT * FROM transactions WHERE 1=1"
        params = []
        if start_date:
            query += " AND date >= ?"
            params.append(start_date)
        if end_date:
            query += " AND date <= ?"
            params.append(end_date)
        query += " ORDER BY date DESC, id DESC"

        rows = conn.execute(query, params).fetchall()

        if own_conn:
            close_connection(conn)

        # ---- 构建 Excel ----
        wb = Workbook()
        ws = wb.active
        ws.title = "交易记录"

        # 样式定义
        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        body_font = Font(name="微软雅黑", size=10)
        body_alignment = Alignment(vertical="center")
        amount_alignment = Alignment(horizontal="right", vertical="center")

        income_font = Font(name="微软雅黑", size=10, color="2E7D32")
        expense_font = Font(name="微软雅黑", size=10, color="C62828")

        thin_border = Border(
            left=Side(style="thin", color="444444"),
            right=Side(style="thin", color="444444"),
            top=Side(style="thin", color="444444"),
            bottom=Side(style="thin", color="444444"),
        )

        # 表头
        headers = ["ID", "日期", "类型", "金额", "分类", "属性", "账户", "备注"]
        col_widths = [6, 12, 8, 12, 10, 14, 10, 30]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 数据行
        total_income = 0.0
        total_expense = 0.0

        for row_idx, row in enumerate(rows, 2):
            values = [
                row["id"],
                row["date"],
                ExportService.TYPE_LABELS.get(row["type"], row["type"]),
                row["amount"],
                row["category"],
                ExportService.ATTR_LABELS.get(row["attribute"], row["attribute"]),
                row["account"],
                row["note"] or "",
            ]

            is_income = row["type"] == "income"
            if is_income:
                total_income += row["amount"]
            else:
                total_expense += row["amount"]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = income_font if is_income else expense_font
                cell.border = thin_border

                if col_idx == 4:  # 金额列
                    cell.alignment = amount_alignment
                    cell.number_format = '#,##0.00'
                elif col_idx == 1:  # ID
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = body_alignment

        # 汇总行
        summary_row = len(rows) + 3
        ws.cell(row=summary_row, column=3, value="支出合计:").font = Font(
            name="微软雅黑", size=10, bold=True, color="C62828"
        )
        ws.cell(row=summary_row, column=4, value=total_expense).font = Font(
            name="微软雅黑", size=10, bold=True, color="C62828"
        )
        ws.cell(row=summary_row, column=4).number_format = '#,##0.00'
        ws.cell(row=summary_row, column=4).alignment = amount_alignment

        summary_row2 = len(rows) + 4
        ws.cell(row=summary_row2, column=3, value="收入合计:").font = Font(
            name="微软雅黑", size=10, bold=True, color="2E7D32"
        )
        ws.cell(row=summary_row2, column=4, value=total_income).font = Font(
            name="微软雅黑", size=10, bold=True, color="2E7D32"
        )
        ws.cell(row=summary_row2, column=4).number_format = '#,##0.00'
        ws.cell(row=summary_row2, column=4).alignment = amount_alignment

        summary_row3 = len(rows) + 5
        balance = total_income - total_expense
        balance_color = "2E7D32" if balance >= 0 else "C62828"
        ws.cell(row=summary_row3, column=3, value="结余:").font = Font(
            name="微软雅黑", size=11, bold=True
        )
        ws.cell(row=summary_row3, column=4, value=balance).font = Font(
            name="微软雅黑", size=11, bold=True, color=balance_color
        )
        ws.cell(row=summary_row3, column=4).number_format = '#,##0.00'
        ws.cell(row=summary_row3, column=4).alignment = amount_alignment

        # 冻结首行
        ws.freeze_panes = "A2"
        # 自动筛选
        ws.auto_filter.ref = f"A1:H{len(rows) + 1}"

        wb.save(filepath)
        return len(rows)
